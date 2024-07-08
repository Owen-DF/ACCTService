from tkinter import *
import ttkbootstrap as ttk
import os
import pandas as pd
import openpyxl
import warnings
import glob
import threading

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)

# File paths for reference, change here for future file reconstructing
YEARLY_COSTING = r"C:\Users\ofleischer\OneDrive - MetroTek Electrical\Coding\SampleFiles\Yearly Costing.xlsx"
BLANK_CELL = 'G101'
OPENPOS = r"C:\Users\ofleischer\OneDrive - MetroTek Electrical\Coding\SampleFiles\Open POs.xlsx"
JOBHOURS = r"C:\Users\ofleischer\OneDrive - MetroTek Electrical\Coding\SampleFiles\Job Hours.xlsx"

# Functions
def submit():
    print("Beginning Process...")
    directory = DInput.get().strip()
    items = IInput.get().strip()
    itemList = filterProjects(items)
    if not itemList:
        print("No valid project codes provided.")
        return

    threading.Thread(target=processItems, args=(directory, itemList)).start()

def filterProjects(items):
    itemList = [item.strip() for item in items.split(",") if item.strip()]
    itemList = [item for item in itemList if len(item) == 4]
    uniqueList = remove_duplicates(itemList)
    return uniqueList

def remove_duplicates(item_list):
    seen = set()
    unique_list = []
    for item in item_list:
        if item not in seen:
            unique_list.append(item)
            seen.add(item)
    return unique_list

def processItems(directory, itemList):
    for item in itemList:
        fileNav(directory, item)
    print("All files processed")

def fileNav(filePath, item):
    matching_folders = []
    pattern = os.path.join(filePath, f"{item}*")
    for folder in glob.glob(pattern):
        if os.path.isdir(folder):
            folder = os.path.join(folder, '03 Estimate & Proposal')
            if os.path.isdir(folder):
                matching_folders.append(folder)

    pullExcel(matching_folders, item)

def pullExcel(matchingFolders, item):
    ExcelFilePaths = []
    if not matchingFolders:
        print(f"No matching folders for item: {item}")
        return

    for folder in matchingFolders:
        pattern = os.path.join(folder, '*.xlsm')
        excelFiles = glob.glob(pattern)
        if excelFiles:
            mostRecent = max(excelFiles, key=os.path.getatime)
            ExcelFilePaths.append(mostRecent)

    PreProcessExcelFiles(ExcelFilePaths, item)

def PreProcessExcelFiles(ExcelList, item):
    df = pd.concat([pd.read_excel(YEARLY_COSTING, sheet_name=sheet_name) for sheet_name in ['2024', '2023', '2022', '2021', '2020', '2019', '2018']])
    df.fillna('', inplace=True)

    df_open_pos = pd.read_excel(OPENPOS)
    df_job_hours = pd.read_excel(JOBHOURS)

    for file_path in ExcelList:
        processExcel(file_path, item, df, df_open_pos, df_job_hours)
    
    print(f"All files processed for item: {item}")

def processExcel(filePath, searchTerm, df, df_open_pos, df_job_hours):
    items = {
        '01.01': 'G9', '01.02': 'G10', '02.01': 'G21', '02.02': 'G22', '02.03': 'G23',
        '03.01': 'G37', '03.02': 'G38', '03.03': 'G39', '04.01': 'G48', '04.02': 'G49',
        '04.03': 'G50', '04.04': 'G51', '05.01': 'G63', '05.02': 'G64', '05.03': 'G65',
        '05.04': 'G66', '06.01': 'G78', '06.02': 'G79', '06.03': 'G80', '06.04': 'G81',
        '06.05': 'G82', '07.0': 'G90', '08.01': 'G94', '08.02': 'G95', '08.03': 'G96', '08.04': 'G97'
    }

    workbook = openpyxl.load_workbook(filePath, read_only=False, keep_vba=True)
    worksheet = workbook['JC']

    for item, cell in items.items():
        result = df.loc[(df['Name'].astype(str).str.contains(searchTerm)) & (df['Item'].astype(str).str.contains(item)), 'Amount'].sum()
        worksheet[cell] = result
    workbook.save(filePath)

    result = df.loc[(df['Name'].astype(str).str.contains(searchTerm)) & (df['Item'] == ''), 'Amount'].sum()
    worksheet[BLANK_CELL] = result
    workbook.save(filePath)

    for item, cell in items.items():
        result = df_open_pos.loc[(df_open_pos['Name'].astype(str).str.contains(searchTerm)) & (df_open_pos['Item'].astype(str).str.contains(item)), 'Open Balance'].sum()
        worksheet[cell.replace('G', 'I')] = result
    workbook.save(filePath)

    totalAmount = df_job_hours.loc[df_job_hours['Job Description'].astype(str).str.contains(searchTerm), 'Hours'].sum()
    worksheet['G87'] = totalAmount
    workbook.save(filePath)
    
    print(f"Processed {searchTerm}")

# GUI setup
app = ttk.Window(themename="superhero")
app.title("EasyAccounting")
app.geometry("600x500")

label = ttk.Label(app, text="EasyAccounting")
label.pack(pady=30)
label.config(font=("Times New Roman", 20, "bold"))

path_frame = ttk.Frame(app)
path_frame.pack(pady=15, padx=10, fill="x")
ttk.Label(path_frame, text="Full Directory").pack(side="left", padx=5)
DInput = ttk.Entry(path_frame)
DInput.pack(side="left", fill="x", expand=True, padx=5)

codes_frame = ttk.Frame(app)
codes_frame.pack(pady=15, padx=10, fill="x")
ttk.Label(codes_frame, text="Project Codes").pack(side="left", padx=5)
IInput = ttk.Entry(codes_frame)
IInput.pack(side="left", fill="x", expand=True, padx=5)

button_frame = ttk.Frame(app)
button_frame.pack(pady=50, padx=10, fill="x")
ttk.Button(button_frame, text="Submit", command=submit, bootstyle="success").pack(side="left", padx=10)

app.mainloop()



# from tkinter import *
# import ttkbootstrap as ttk
# import os
# import pandas as pd
# import openpyxl
# import warnings
# import glob
# import threading
# from zipfile import BadZipFile

# warnings.filterwarnings("ignore", category=UserWarning)
# warnings.filterwarnings("ignore", category=FutureWarning)

# # File paths for reference, change here for future file reconstructing
# YEARLY_COSTING = r"C:\Users\ofleischer\OneDrive - MetroTek Electrical\Coding\SampleFiles\Yearly Costing.xlsx"
# BLANK_CELL = 'G101'
# OPENPOS = r'C:\Users\ofleischer\OneDrive - MetroTek Electrical\Coding\SampleFiles\Open POs.xlsx'
# JOBHOURS = r'C:\Users\ofleischer\OneDrive - MetroTek Electrical\Coding\SampleFiles\Job Hours.xlsx'

# # Functions
# def submit():
#     directory = DInput.get().strip()
#     bar['value'] = 0
#     items = IInput.get().strip()
#     itemList = filterProjects(items)
    
#     threading.Thread(target=fileNav, args=(directory, itemList)).start()
#     return

# def filterProjects(items):
#     itemList = [item.strip() for item in items.split(",") if item.strip()]
#     itemList = [item for item in itemList if len(item) == 4]
#     uniqueList = remove_duplicates(itemList)
#     return uniqueList

# def remove_duplicates(item_list):
#     seen = set()
#     unique_list = []
#     for item in item_list:
#         if item not in seen:
#             unique_list.append(item)
#             seen.add(item)
#     return unique_list

# def fileNav(filePath, itemList):
#     matching_folders = []
#     for item in itemList:
#         pattern = os.path.join(filePath, f"{item}*")
#         for folder in glob.glob(pattern):
#             if os.path.isdir(folder):
#                 folder = os.path.join(folder, '03 Estimate & Proposal')
#                 if os.path.isdir(folder):
#                     matching_folders.append(folder)
    
#     pullExcel(matching_folders, itemList)
#     return

# def pullExcel(matchingFolders, itemList):
#     ExcelFilePaths = []
#     if not matchingFolders:
#         return
#     try:
#         stepValue = 100 / len(matchingFolders)
#     except ZeroDivisionError:
#         return
#     for folder in matchingFolders:
#         excelFiles = []
#         pattern = os.path.join(folder, '*.xlsm')
#         for file in glob.glob(pattern):
#             excelFiles.append(file)
#         if excelFiles:
#             if len(excelFiles) > 1:
#                 mostRecent = max(excelFiles, key=os.path.getatime)
#                 ExcelFilePaths.append(mostRecent)
#             else:
#                 ExcelFilePaths.append(excelFiles[0])
#         bar['value'] += stepValue
#         app.update_idletasks()

#     PreProcessExcelFiles(ExcelFilePaths, itemList)

# def PreProcessExcelFiles(ExcelList, itemList):
#     try:
#         df = pd.concat([pd.read_excel(YEARLY_COSTING, sheet_name=sheet_name) for sheet_name in ['2024', '2023', '2022', '2021', '2020', '2019', '2018']])
#     except FileNotFoundError:
#         print(f"File not found: {YEARLY_COSTING}")
#         return
#     df.fillna('', inplace=True)

#     for i, file_path in enumerate(ExcelList):
#         searchTerm = itemList[i]
#         processExcel(file_path, searchTerm, df)

# def processExcel(filePath, searchTerm, df):
#     if not os.path.exists(filePath):
#         print(f"File not found: {filePath}")
#         return

#     items = {
#         '01.01': 'G9', '01.02': 'G10', '02.01': 'G21', '02.02': 'G22', '02.03': 'G23',
#         '03.01': 'G37', '03.02': 'G38', '03.03': 'G39', '04.01': 'G48', '04.02': 'G49',
#         '04.03': 'G50', '04.04': 'G51', '05.01': 'G63', '05.02': 'G64', '05.03': 'G65',
#         '05.04': 'G66', '06.01': 'G78', '06.02': 'G79', '06.03': 'G80', '06.04': 'G81',
#         '06.05': 'G82', '07.0': 'G90', '08.01': 'G94', '08.02': 'G95', '08.03': 'G96', '08.04': 'G97'
#     }

#     try:
#         workbook = openpyxl.load_workbook(filePath, read_only=False, keep_vba=True)
#     except BadZipFile:
#         print(f"Bad zip file: {filePath}")
#         return

#     worksheet = workbook['JC']

#     for item, cell in items.items():
#         result = df.loc[(df['Name'].astype(str).str.contains(searchTerm)) & (df['Item'].astype(str).str.contains(item)), 'Amount'].sum()
#         worksheet[cell] = result
#     workbook.save(filePath)

#     result = df.loc[(df['Name'].astype(str).str.contains(searchTerm)) & (df['Item'] == ''), 'Amount'].sum()
#     cell = BLANK_CELL
#     worksheet[cell] = result
#     workbook.save(filePath)

#     try:
#         df_open_pos = pd.read_excel(OPENPOS)
#     except FileNotFoundError:
#         print(f"File not found: {OPENPOS}")
#         return

#     items = {
#         '01.01': 'I9', '01.02': 'I10', '02.01': 'I21', '02.02': 'I22', '02.03': 'I23',
#         '03.01': 'I37', '03.02': 'I38', '03.03': 'I39', '04.01': 'I48', '04.02': 'I49',
#         '04.03': 'I50', '04.04': 'I51', '05.01': 'I63', '05.02': 'I64', '05.03': 'I65',
#         '05.04': 'I66', '06.01': 'I78', '06.02': 'I79', '06.03': 'I80', '06.04': 'I81',
#         '06.05': 'I82', '07.0': 'I90', '08.01': 'I94', '08.02': 'I95', '08.03': 'I96', '08.04': 'I97'
#     }

#     workbook = openpyxl.load_workbook(filePath, read_only=False, keep_vba=True)
#     worksheet = workbook['JC']

#     for item, cell in items.items():
#         result = df_open_pos.loc[(df_open_pos['Name'].astype(str).str.contains(searchTerm)) & (df_open_pos['Item'].astype(str).str.contains(item)), 'Open Balance'].sum()
#         worksheet[cell] = result
#     workbook.save(filePath)

#     try:
#         dfJobHours = pd.read_excel(JOBHOURS)
#     except FileNotFoundError:
#         print(f"File not found: {JOBHOURS}")
#         return

#     totalAmount = dfJobHours.loc[dfJobHours['Job Description'].astype(str).str.contains(searchTerm), 'Hours'].sum()

#     cell = 'G87'
#     worksheet[cell] = totalAmount
#     workbook.save(filePath)
#     return

# # GUI setup
# app = ttk.Window(themename="superhero")
# app.title("EasyAccounting")
# app.geometry("600x500")

# label = ttk.Label(app, text="EasyAccounting")
# label.pack(pady=30)
# label.config(font=("Times New Roman", 20, "bold"))

# path_frame = ttk.Frame(app)
# path_frame.pack(pady=15, padx=10, fill="x")
# ttk.Label(path_frame, text="Full Directory").pack(side="left", padx=5)
# DInput = ttk.Entry(path_frame)
# DInput.pack(side="left", fill="x", expand=True, padx=5)

# codes_frame = ttk.Frame(app)
# codes_frame.pack(pady=15, padx=10, fill="x")
# ttk.Label(codes_frame, text="Project Codes").pack(side="left", padx=5)
# IInput = ttk.Entry(codes_frame)
# IInput.pack(side="left", fill="x", expand=True, padx=5)

# button_frame = ttk.Frame(app)
# button_frame.pack(pady=50, padx=10, fill="x")
# ttk.Button(button_frame, text="Submit", command=submit, bootstyle="success").pack(side="left", padx=10)

# progress_frame = ttk.Frame(app)
# progress_frame.pack(pady=25, padx=10, fill="x")
# bar = ttk.Progressbar(progress_frame, length=400, style='success.Striped.Horizontal.TProgressbar')
# bar.pack(pady=10)

# app.mainloop()

